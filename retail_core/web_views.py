"""
Web Views for Retail Analytics Platform
Django template views for the web interface
"""

from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.urls import reverse_lazy
from django.shortcuts import redirect
from django.db.models import Sum, Count, Avg, Q, F, Max, DecimalField
from django.db.models.functions import TruncDate, TruncMonth
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
from datetime import timedelta
from decimal import Decimal
from products.models import Product, Category, Supplier
from inventory.models import InventoryLevel, Store
from orders.models import Order, Customer
from analytics.models import DailySalesMetrics
import json
import csv
from io import StringIO, BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


# ============ Product Views ============

class ProductListView(ListView):
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'
    paginate_by = 20

    def get_queryset(self):
        queryset = Product.objects.select_related('category', 'supplier').prefetch_related('inventory_levels')
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | 
                Q(sku__icontains=search) | 
                Q(description__icontains=search)
            )
        
        # Category filter
        category = self.request.GET.get('category')
        if category:
            queryset = queryset.filter(category_id=category)
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.all()
        
        # Add total stock for each product
        for product in context['products']:
            product.total_stock = product.inventory_levels.aggregate(
                total=Sum('quantity_on_hand')
            )['total'] or 0
        
        return context


class ProductDetailView(DetailView):
    model = Product
    template_name = 'products/product_detail.html'
    context_object_name = 'product'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        product = self.get_object()
        
        # Inventory levels
        context['inventory_levels'] = InventoryLevel.objects.filter(
            product=product
        ).select_related('store')
        
        # Calculate totals
        context['total_stock'] = context['inventory_levels'].aggregate(
            total=Sum('quantity_on_hand')
        )['total'] or 0
        
        context['total_value'] = context['total_stock'] * product.cost_price
        
        # Sales last 30 days
        thirty_days_ago = timezone.now() - timedelta(days=30)
        context['sales_30days'] = Order.objects.filter(
            line_items__product=product,
            order_date__gte=thirty_days_ago
        ).aggregate(total=Sum('line_items__quantity'))['total'] or 0
        
        # Recent activity (mock data for now)
        context['recent_activities'] = []
        
        return context


class ProductCreateView(CreateView):
    model = Product
    template_name = 'products/product_form.html'
    fields = ['name', 'sku', 'barcode', 'category', 'supplier', 'description',
              'cost_price', 'selling_price', 'reorder_point', 'reorder_quantity', 
              'max_stock', 'is_active']
    success_url = reverse_lazy('product_list')

    def form_valid(self, form):
        messages.success(self.request, 'Product created successfully!')
        return super().form_valid(form)


class ProductUpdateView(UpdateView):
    model = Product
    template_name = 'products/product_form.html'
    fields = ['name', 'sku', 'barcode', 'category', 'supplier', 'description',
              'cost_price', 'selling_price', 'reorder_point', 'reorder_quantity', 
              'max_stock', 'is_active']
    success_url = reverse_lazy('product_list')

    def form_valid(self, form):
        messages.success(self.request, 'Product updated successfully!')
        return super().form_valid(form)


class ProductDeleteView(DeleteView):
    model = Product
    success_url = reverse_lazy('product_list')

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Product deleted successfully!')
        return super().delete(request, *args, **kwargs)


# ============ Inventory Views ============

class InventoryListView(ListView):
    model = InventoryLevel
    template_name = 'inventory/inventory_list.html'
    context_object_name = 'inventory_items'
    paginate_by = 50

    def get_queryset(self):
        queryset = InventoryLevel.objects.select_related('product', 'store')
        
        # Store filter
        store = self.request.GET.get('store')
        if store:
            queryset = queryset.filter(store_id=store)
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'low':
            queryset = queryset.filter(quantity_on_hand__lte=F('product__reorder_point'))
        elif status == 'out':
            queryset = queryset.filter(quantity_on_hand=0)
        elif status == 'over':
            queryset = queryset.filter(quantity_on_hand__gte=F('product__max_stock'))
        
        # Search
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(product__name__icontains=search) | 
                Q(product__sku__icontains=search)
            )
        
        return queryset.order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['stores'] = Store.objects.filter(is_active=True)
        
        # Calculate stats
        all_inventory = InventoryLevel.objects.all()
        context['total_items'] = all_inventory.aggregate(
            total=Sum('quantity_on_hand')
        )['total'] or 0
        
        context['total_value'] = sum([
            (item.quantity_on_hand * item.product.cost_price) 
            for item in all_inventory.select_related('product')
        ])
        
        context['low_stock_count'] = all_inventory.filter(
            quantity_on_hand__lte=F('product__reorder_point')
        ).count()
        
        context['out_of_stock_count'] = all_inventory.filter(quantity_on_hand=0).count()
        
        return context


class InventoryDetailView(DetailView):
    model = InventoryLevel
    template_name = 'inventory/inventory_detail.html'
    context_object_name = 'inventory'


# ============ Order Views ============

class OrderListView(ListView):
    model = Order
    template_name = 'orders/order_list.html'
    context_object_name = 'orders'
    paginate_by = 25

    def get_queryset(self):
        queryset = Order.objects.select_related('customer', 'store').prefetch_related('line_items')
        
        # Status filter
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Date filters
        from_date = self.request.GET.get('from_date')
        to_date = self.request.GET.get('to_date')
        if from_date:
            queryset = queryset.filter(order_date__gte=from_date)
        if to_date:
            queryset = queryset.filter(order_date__lte=to_date)
        
        return queryset.order_by('-order_date')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = timezone.now().date()
        
        # Today's stats
        context['today_orders'] = Order.objects.filter(order_date=today).count()
        context['today_revenue'] = Order.objects.filter(order_date=today).aggregate(
            total=Sum('total')
        )['total'] or 0
        
        # Status counts
        context['pending_orders'] = Order.objects.filter(status='pending').count()
        context['processing_orders'] = Order.objects.filter(status='processing').count()
        
        return context


class OrderDetailView(DetailView):
    model = Order
    template_name = 'orders/order_detail.html'
    context_object_name = 'order'


# ============ Analytics Views ============

class AnalyticsDashboardView(TemplateView):
    template_name = 'analytics/analytics_dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Date ranges
        today = timezone.now().date()
        thirty_days_ago = today - timedelta(days=30)
        last_month_start = thirty_days_ago - timedelta(days=30)
        
        # Total revenue (30 days)
        current_revenue = Order.objects.filter(
            order_date__gte=thirty_days_ago,
            status__in=['confirmed', 'shipped', 'delivered']
        ).aggregate(total=Sum('total'))['total'] or 0
        
        last_month_revenue = Order.objects.filter(
            order_date__gte=last_month_start,
            order_date__lt=thirty_days_ago,
            status__in=['confirmed', 'shipped', 'delivered']
        ).aggregate(total=Sum('total'))['total'] or 1
        
        context['total_revenue'] = f"{current_revenue:,.2f}"
        context['revenue_growth'] = f"{((current_revenue - last_month_revenue) / last_month_revenue * 100):.1f}"
        
        # Total orders
        current_orders = Order.objects.filter(order_date__gte=thirty_days_ago).count()
        last_month_orders = Order.objects.filter(
            order_date__gte=last_month_start,
            order_date__lt=thirty_days_ago
        ).count() or 1
        
        context['total_orders'] = current_orders
        context['orders_growth'] = f"{((current_orders - last_month_orders) / last_month_orders * 100):.1f}"
        
        # Average order value
        context['avg_order_value'] = f"{(current_revenue / current_orders if current_orders > 0 else 0):.2f}"
        
        # Active products
        context['active_products'] = Product.objects.filter(is_active=True).count()
        
        # Sales trend data (last 30 days)
        sales_data = []
        sales_dates = []
        for i in range(30):
            date = today - timedelta(days=29-i)
            sales_dates.append(date.strftime('%m/%d'))
            daily_sales = Order.objects.filter(
                order_date=date,
                status__in=['confirmed', 'shipped', 'delivered']
            ).aggregate(total=Sum('total'))['total'] or 0
            sales_data.append(float(daily_sales))
        
        context['sales_dates'] = sales_dates
        context['sales_data'] = sales_data
        
        # Top categories
        from django.db.models import Sum as DbSum
        category_sales = Order.objects.filter(
            order_date__gte=thirty_days_ago
        ).values('line_items__product__category__name').annotate(
            total=DbSum('line_items__quantity')
        ).order_by('-total')[:5]
        
        context['category_labels'] = [cat['line_items__product__category__name'] or 'Uncategorized' 
                                      for cat in category_sales]
        context['category_data'] = [int(cat['total']) for cat in category_sales]
        
        # Top products
        context['top_products'] = []
        
        # Top customers
        context['top_customers'] = []
        
        # Inventory health
        total_inventory = InventoryLevel.objects.count()
        low_stock = InventoryLevel.objects.filter(
            quantity_on_hand__lte=F('product__reorder_point'),
            quantity_on_hand__gt=0
        ).count()
        out_of_stock = InventoryLevel.objects.filter(quantity_on_hand=0).count()
        overstock = InventoryLevel.objects.filter(
            quantity_on_hand__gte=F('product__max_stock')
        ).count()
        in_stock = total_inventory - low_stock - out_of_stock - overstock
        
        context['inventory_health'] = [in_stock, low_stock, out_of_stock, overstock]
        
        return context


class ReportsView(TemplateView):
    """Reports hub - displays all available report categories"""
    template_name = 'reports/reports.html'


# ============ Sales Reports ============

class DailySalesReportView(LoginRequiredMixin, TemplateView):
    """Daily sales report with daily breakdown"""
    template_name = 'reports/sales/daily_sales.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get last 30 days of sales
        thirty_days_ago = timezone.now() - timedelta(days=30)
        daily_sales = Order.objects.filter(
            order_date__gte=thirty_days_ago,
            status__in=['CONFIRMED', 'SHIPPED', 'DELIVERED']
        ).values('order_date__date').annotate(
            revenue=Sum('total'),
            order_count=Count('id'),
            avg_order_value=Avg('total')
        ).order_by('order_date__date')
        
        context['daily_sales'] = list(daily_sales)
        context['total_revenue'] = sum(item['revenue'] or 0 for item in daily_sales)
        context['total_orders'] = sum(item['order_count'] or 0 for item in daily_sales)
        context['avg_daily_revenue'] = context['total_revenue'] / 30 if context['total_revenue'] else 0
        
        return context


class MonthlySalesReportView(LoginRequiredMixin, TemplateView):
    """Monthly sales report with monthly trends"""
    template_name = 'reports/sales/monthly_sales.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get last 12 months of sales
        one_year_ago = timezone.now() - timedelta(days=365)
        monthly_sales = Order.objects.filter(
            order_date__gte=one_year_ago,
            status__in=['CONFIRMED', 'SHIPPED', 'DELIVERED']
        ).values('order_date__year', 'order_date__month').annotate(
            revenue=Sum('total'),
            order_count=Count('id'),
            avg_order_value=Avg('total')
        ).order_by('order_date__year', 'order_date__month')
        
        context['monthly_sales'] = list(monthly_sales)
        context['total_revenue'] = sum(item['revenue'] or 0 for item in monthly_sales)
        context['total_orders'] = sum(item['order_count'] or 0 for item in monthly_sales)
        
        return context


class SalesByProductReportView(LoginRequiredMixin, TemplateView):
    """Sales by product report with top sellers"""
    template_name = 'reports/sales/sales_by_product.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get top 20 products by revenue
        product_sales = Order.objects.filter(
            status__in=['CONFIRMED', 'SHIPPED', 'DELIVERED']
        ).values('line_items__product__name', 'line_items__product__sku').annotate(
            units_sold=Sum('line_items__quantity'),
            revenue=Sum('line_items__line_total'),
            avg_price=Avg('line_items__unit_price')
        ).order_by('-revenue')[:20]
        
        context['product_sales'] = list(product_sales)
        context['total_revenue'] = sum(item['revenue'] or 0 for item in product_sales)
        context['total_units'] = sum(item['units_sold'] or 0 for item in product_sales)
        
        return context


class SalesByCategoryReportView(LoginRequiredMixin, TemplateView):
    """Sales by category report"""
    template_name = 'reports/sales/sales_by_category.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get sales by category
        category_sales = Order.objects.filter(
            status__in=['CONFIRMED', 'SHIPPED', 'DELIVERED']
        ).values('line_items__product__category__name').annotate(
            units_sold=Sum('line_items__quantity'),
            revenue=Sum('line_items__line_total')
        ).order_by('-revenue')
        
        context['category_sales'] = list(category_sales)
        context['total_revenue'] = sum(item['revenue'] or 0 for item in category_sales)
        
        return context


# ============ Inventory Reports ============

class StockLevelsReportView(LoginRequiredMixin, TemplateView):
    """Current stock levels report"""
    template_name = 'reports/inventory/stock_levels.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get current stock levels
        stock_levels = InventoryLevel.objects.select_related(
            'product', 'store'
        ).values(
            'product__name', 'product__sku', 'store__name', 
            'quantity_on_hand', 'product__reorder_point'
        ).order_by('product__name')
        
        context['stock_levels'] = list(stock_levels)
        context['total_items'] = len(stock_levels)
        context['total_value'] = sum(
            item['quantity_on_hand'] * 50 for item in stock_levels  # Rough estimate
        )
        
        return context


class LowStockAlertReportView(LoginRequiredMixin, TemplateView):
    """Low stock alert report"""
    template_name = 'reports/inventory/low_stock.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get low stock items
        low_stock = InventoryLevel.objects.filter(
            quantity_on_hand__lte=F('product__reorder_point'),
            quantity_on_hand__gt=0
        ).select_related('product', 'store').values(
            'product__name', 'product__sku', 'store__name',
            'quantity_on_hand', 'product__reorder_point', 'product__cost_price'
        ).order_by('quantity_on_hand')
        
        context['low_stock_items'] = list(low_stock)
        context['critical_count'] = len(low_stock)
        
        return context


class StockMovementReportView(LoginRequiredMixin, TemplateView):
    """Stock movement report"""
    template_name = 'reports/inventory/stock_movement.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get recent stock movements (orders)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        movements = Order.objects.filter(
            order_date__gte=thirty_days_ago
        ).values('line_items__product__name', 'line_items__product__sku').annotate(
            movement_qty=Sum('line_items__quantity'),
            movement_date=Max('order_date')
        ).order_by('-movement_qty')[:20]
        
        context['movements'] = list(movements)
        
        return context


class InventoryValuationReportView(LoginRequiredMixin, TemplateView):
    """Inventory valuation report"""
    template_name = 'reports/inventory/inventory_valuation.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get inventory valuation by store
        valuation = InventoryLevel.objects.select_related(
            'product', 'store'
        ).values('store__name').annotate(
            total_units=Sum('quantity_on_hand'),
            total_value=Sum(F('quantity_on_hand') * F('product__cost_price'), output_field=DecimalField())
        ).order_by('-total_value')
        
        context['valuation'] = list(valuation)
        context['total_inventory_value'] = sum(item['total_value'] or 0 for item in valuation)
        
        return context


# ============ Customer Reports ============

class CustomerPurchasesReportView(LoginRequiredMixin, TemplateView):
    """Customer purchases report"""
    template_name = 'reports/customer/customer_purchases.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get customer purchase history
        customer_purchases = Order.objects.values('customer__name', 'customer__email').annotate(
            purchase_count=Count('id'),
            total_spent=Sum('total'),
            avg_order_value=Avg('total'),
            last_purchase=Max('order_date')
        ).order_by('-total_spent')[:50]
        
        context['customer_purchases'] = list(customer_purchases)
        context['total_customers'] = len(customer_purchases)
        
        return context


class TopCustomersReportView(LoginRequiredMixin, TemplateView):
    """Top customers report"""
    template_name = 'reports/customer/top_customers.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get top 20 customers by revenue
        top_customers = Order.objects.values(
            'customer__name', 'customer__email', 'customer__loyalty_points'
        ).annotate(
            purchase_count=Count('id'),
            total_spent=Sum('total'),
            avg_order_value=Avg('total')
        ).order_by('-total_spent')[:20]
        
        context['top_customers'] = list(top_customers)
        context['vip_threshold'] = 5000  # $5000 threshold for VIP
        
        return context


class CustomerRetentionReportView(LoginRequiredMixin, TemplateView):
    """Customer retention report"""
    template_name = 'reports/customer/customer_retention.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate retention metrics
        thirty_days_ago = timezone.now() - timedelta(days=30)
        sixty_days_ago = timezone.now() - timedelta(days=60)
        ninety_days_ago = timezone.now() - timedelta(days=90)
        
        recent_customers = Order.objects.filter(
            order_date__gte=thirty_days_ago
        ).values('customer_id').distinct().count()
        
        recurring_customers = Order.objects.filter(
            order_date__gte=thirty_days_ago
        ).values('customer_id').annotate(
            order_count=Count('id')
        ).filter(order_count__gt=1).count()
        
        context['recent_customers'] = recent_customers
        context['recurring_customers'] = recurring_customers
        retention_rate = (recurring_customers / recent_customers * 100) if recent_customers > 0 else 0
        context['retention_rate'] = retention_rate
        context['churn_rate'] = 100 - retention_rate
        
        return context


# ============ Performance Reports ============

class ProfitMarginReportView(LoginRequiredMixin, TemplateView):
    """Profit margin analysis report"""
    template_name = 'reports/performance/profit_margin.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get profit margins by product
        from django.db.models import F, DecimalField
        profit_data = Order.objects.filter(
            status__in=['CONFIRMED', 'SHIPPED', 'DELIVERED']
        ).values('line_items__product__name').annotate(
            revenue=Sum('line_items__line_total'),
            cost=Sum(F('line_items__quantity') * F('line_items__unit_price') * Decimal('0.7'), 
                    output_field=DecimalField())
        ).order_by('-revenue')[:20]
        
        for item in profit_data:
            item['profit'] = (item['revenue'] or 0) - (item['cost'] or 0)
            item['margin_percent'] = ((item['profit'] / (item['revenue'] or 1)) * 100) if item['revenue'] else 0
        
        context['profit_data'] = list(profit_data)
        
        return context


class SupplierPerformanceReportView(LoginRequiredMixin, TemplateView):
    """Supplier performance report"""
    template_name = 'reports/performance/supplier_performance.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get supplier performance metrics
        supplier_perf = Product.objects.values('supplier__name').annotate(
            product_count=Count('id'),
            total_revenue=Sum('order_lines__order__total')
        ).order_by('-total_revenue')
        
        context['supplier_performance'] = list(supplier_perf)
        
        return context


class ForecastAccuracyReportView(LoginRequiredMixin, TemplateView):
    """Forecast accuracy report"""
    template_name = 'reports/performance/forecast_accuracy.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Mock forecast accuracy data
        context['forecast_metrics'] = {
            'mae': 12.5,
            'rmse': 15.3,
            'mape': 8.2,
            'accuracy': 91.8,
            'models': ['ARIMA', 'Prophet', 'LSTM', 'Ensemble']
        }
        
        context['forecast_performance'] = [
            {'product': 'Premium Coffee Beans', 'predicted': 245, 'actual': 238, 'error_percent': 2.9},
            {'product': 'Organic Milk', 'predicted': 189, 'actual': 195, 'error_percent': 3.1},
            {'product': 'Whole Wheat Bread', 'predicted': 156, 'actual': 152, 'error_percent': 2.6},
            {'product': 'Fresh Tomatoes', 'predicted': 134, 'actual': 140, 'error_percent': 4.3},
            {'product': 'Free Range Eggs', 'predicted': 123, 'actual': 118, 'error_percent': 4.1},
        ]
        
        return context

class ForecastingDashboardView(TemplateView):
    """Demand forecasting dashboard with predictions"""
    template_name = 'forecasting/forecasting.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Get top products for forecasting
        top_products = Product.objects.filter(is_active=True).order_by('-created_at')[:10]
        context['products'] = top_products
        
        # Generate forecast data for next 30 days
        today = timezone.now().date()
        forecasts = []
        
        for days_ahead in range(1, 31):
            forecast_date = today + timedelta(days=days_ahead)
            # Generate realistic demand forecast (normally would come from ML model)
            base_demand = 100
            trend = days_ahead * 0.5
            seasonal = 20 * (1 + (days_ahead % 7) / 7)
            forecast = max(50, base_demand + trend + seasonal)
            
            forecasts.append({
                'date': forecast_date.isoformat(),
                'forecast': round(forecast, 2),
                'confidence': round(85 + (days_ahead % 10) * 0.5, 1),  # 85-90% confidence
            })
        
        context['forecasts'] = forecasts
        
        # Forecast accuracy metrics
        context['forecast_accuracy'] = {
            'mae': 12.5,  # Mean Absolute Error
            'rmse': 15.3,  # Root Mean Squared Error
            'mape': 8.2,  # Mean Absolute Percentage Error
        }
        
        # Top predicted demand products
        context['top_forecast_products'] = [
            {'name': p.name, 'sku': p.sku, 'forecast': round(150 + (i * 20)), 'confidence': 87.5}
            for i, p in enumerate(top_products[:5])
        ]
        
        # Forecast alerts
        context['forecast_alerts'] = [
            {'type': 'high', 'message': 'High demand predicted for Laptop Pro in next 7 days', 'action': 'Increase inventory'},
            {'type': 'medium', 'message': 'Seasonal spike expected for Electronics category', 'action': 'Prepare for surge'},
            {'type': 'low', 'message': 'Declining demand for Clothing items', 'action': 'Plan promotions'},
        ]
        
        return context


# ============ Shelf Vision Views ============

class ShelfVisionDashboardView(TemplateView):
    """Computer vision shelf analysis dashboard"""
    template_name = 'shelf_vision/shelf_vision.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Store list for camera feeds
        context['stores'] = Store.objects.filter(is_active=True)
        
        # Shelf analysis data
        shelves = [
            {
                'id': 1,
                'store': 'Main Store',
                'location': 'Aisle A-1',
                'product': 'Laptop Pro',
                'facing': 5,
                'stock_level': 'Good',
                'compliance': 95,
                'last_scan': '2026-02-07 14:30:00'
            },
            {
                'id': 2,
                'store': 'Main Store',
                'location': 'Aisle B-2',
                'product': 'Wireless Mouse',
                'facing': 3,
                'stock_level': 'Low',
                'compliance': 65,
                'last_scan': '2026-02-07 14:25:00'
            },
            {
                'id': 3,
                'store': 'North Branch',
                'location': 'Aisle C-1',
                'product': 'USB Cable',
                'facing': 8,
                'stock_level': 'Excellent',
                'compliance': 100,
                'last_scan': '2026-02-07 14:35:00'
            },
        ]
        context['shelves'] = shelves
        
        # Issues detected
        context['issues'] = [
            {'severity': 'high', 'issue': 'Out of stock - Wireless Mouse Aisle B-2', 'action': 'Restock immediately', 'store': 'Main Store'},
            {'severity': 'medium', 'issue': 'Planogram violation - Monitor 27" arrangement incorrect', 'action': 'Fix arrangement', 'store': 'North Branch'},
            {'severity': 'low', 'issue': 'Price tag missing - USB Cable', 'action': 'Update price tag', 'store': 'South Branch'},
        ]
        
        # Camera status
        context['camera_status'] = {
            'total_cameras': 12,
            'active': 11,
            'inactive': 1,
            'last_sync': '2026-02-07 15:00:00'
        }
        
        # Shelf compliance metrics
        context['compliance_metrics'] = {
            'overall_compliance': 87.3,
            'by_category': [
                {'category': 'Electronics', 'compliance': 92},
                {'category': 'Clothing', 'compliance': 85},
                {'category': 'Food & Beverage', 'compliance': 88},
                {'category': 'Home & Garden', 'compliance': 80},
                {'category': 'Sports', 'compliance': 90},
            ]
        }
        
        return context


# ============ Alerts Management Views ============

class AlertsDashboardView(TemplateView):
    """Real-time alerts management dashboard"""
    template_name = 'alerts/alerts.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Active alerts
        alerts = [
            {
                'id': 1,
                'type': 'stock',
                'severity': 'high',
                'title': 'Low Stock Alert',
                'message': '5 products below reorder point',
                'timestamp': '2026-02-07 15:30:00',
                'status': 'new',
                'action_url': '/inventory/?status=low_stock'
            },
            {
                'id': 2,
                'type': 'sales',
                'severity': 'medium',
                'title': 'Unusual Sales Pattern',
                'message': 'Laptop Pro sales up 150% compared to last week',
                'timestamp': '2026-02-07 15:15:00',
                'status': 'acknowledged',
                'action_url': '/orders/?product=laptop_pro'
            },
            {
                'id': 3,
                'type': 'forecast',
                'severity': 'medium',
                'title': 'Forecast Accuracy Low',
                'message': 'Forecast accuracy for Q2 dropped to 72%',
                'timestamp': '2026-02-07 15:00:00',
                'status': 'new',
                'action_url': '/forecasting/'
            },
            {
                'id': 4,
                'type': 'vision',
                'severity': 'high',
                'title': 'Shelf Compliance Issue',
                'message': 'Multiple planogram violations detected in Main Store',
                'timestamp': '2026-02-07 14:45:00',
                'status': 'acknowledged',
                'action_url': '/shelf-vision/'
            },
            {
                'id': 5,
                'type': 'supplier',
                'severity': 'medium',
                'title': 'Supplier Delay',
                'message': 'Supplier B order delayed by 2 days',
                'timestamp': '2026-02-07 14:30:00',
                'status': 'resolved',
                'action_url': '/products/?supplier=supplier_b'
            },
        ]
        context['alerts'] = alerts
        
        # Alert statistics
        context['stats'] = {
            'total': len(alerts),
            'new': sum(1 for a in alerts if a['status'] == 'new'),
            'acknowledged': sum(1 for a in alerts if a['status'] == 'acknowledged'),
            'resolved': sum(1 for a in alerts if a['status'] == 'resolved'),
            'high_severity': sum(1 for a in alerts if a['severity'] == 'high'),
        }
        
        # Alert types breakdown
        context['alert_types'] = [
            {'type': 'Stock Alerts', 'count': 24, 'color': '#dc3545'},
            {'type': 'Sales Alerts', 'count': 18, 'color': '#fd7e14'},
            {'type': 'Forecast Alerts', 'count': 12, 'color': '#0dcaf0'},
            {'type': 'Vision Alerts', 'count': 15, 'color': '#198754'},
            {'type': 'Supplier Alerts', 'count': 8, 'color': '#6f42c1'},
        ]
        
        # Alert thresholds/rules
        context['alert_rules'] = [
            {'rule': 'Stock below reorder point', 'enabled': True, 'severity': 'high'},
            {'rule': 'Out of stock for 24 hours', 'enabled': True, 'severity': 'critical'},
            {'rule': 'Sales variance > 50%', 'enabled': True, 'severity': 'medium'},
            {'rule': 'Forecast accuracy < 75%', 'enabled': True, 'severity': 'medium'},
            {'rule': 'Planogram violation', 'enabled': True, 'severity': 'high'},
            {'rule': 'Supplier delivery late', 'enabled': True, 'severity': 'medium'},
        ]
        
        return context


# ============ Export & Filter API Views ============

class ExportAllReportsView(LoginRequiredMixin, View):
    """Export all reports data to various formats (PDF, Excel, CSV)"""
    
    def _get_sales_data(self, date_from=None, date_to=None):
        """Get sales report data"""
        filters = {}
        if date_from:
            filters['order_date__gte'] = date_from
        if date_to:
            filters['order_date__lte'] = date_to
            
        sales = Order.objects.filter(**filters).values(
            'id', 'order_date', 'customer__name', 'total', 'status'
        ).order_by('-order_date')
        return list(sales), ['Order ID', 'Date', 'Customer', 'Total', 'Status']
    
    def _get_inventory_data(self):
        """Get inventory report data"""
        inventory = InventoryLevel.objects.select_related(
            'product', 'store'
        ).values(
            'product__name', 'product__sku', 'store__name', 
            'quantity_on_hand', 'product__cost_price'
        ).order_by('product__name')
        return list(inventory), ['Product', 'SKU', 'Store', 'Quantity', 'Cost Price']
    
    def _get_customer_data(self):
        """Get customer report data"""
        customers = Customer.objects.values(
            'id', 'name', 'email', 'phone', 'loyalty_points'
        ).order_by('name')
        return list(customers), ['ID', 'Name', 'Email', 'Phone', 'Loyalty Points']
    
    def _generate_csv(self, data, headers):
        """Generate CSV file"""
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        
        for row in data:
            if isinstance(row, dict):
                writer.writerow([row.get(key.lower().replace(' ', '_'), '') for key in headers])
            else:
                writer.writerow(row)
        
        return output.getvalue()
    
    def _generate_excel(self, data, headers):
        """Generate Excel file"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data
        for row_idx, row in enumerate(data, 2):
            if isinstance(row, dict):
                for col_idx, header in enumerate(headers, 1):
                    key = header.lower().replace(' ', '_')
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = row.get(key, '')
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    def _generate_pdf(self, data, headers, title="Report"):
        """Generate PDF file"""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        story = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph(title, title_style))
        story.append(Spacer(1, 0.3*inch))
        
        # Convert data to table format
        table_data = [headers]
        for row in data[:100]:  # Limit to 100 rows for PDF readability
            if isinstance(row, dict):
                table_data.append([str(row.get(key.lower().replace(' ', '_'), ''))[:50] for key in headers])
        
        # Create table
        table = Table(table_data, colWidths=[1.2*inch] * len(headers))
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 0.3*inch))
        
        # Footer
        footer_style = ParagraphStyle(
            'CustomFooter',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        story.append(Paragraph(f"Generated on {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
        
        doc.build(story)
        output.seek(0)
        return output.getvalue()
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            export_format = data.get('format', 'pdf').lower()
            report_type = data.get('report_type', 'all')
            date_from = data.get('date_from')
            date_to = data.get('date_to')
            
            # Validate format
            if export_format not in ['pdf', 'excel', 'csv']:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid format. Use pdf, excel, or csv'
                }, status=400)
            
            # Get report data
            if report_type == 'sales' or report_type == 'all':
                report_data, headers = self._get_sales_data(date_from, date_to)
                title = 'Sales Report'
            elif report_type == 'inventory':
                report_data, headers = self._get_inventory_data()
                title = 'Inventory Report'
            elif report_type == 'customers':
                report_data, headers = self._get_customer_data()
                title = 'Customer Report'
            else:
                report_data, headers = self._get_sales_data(date_from, date_to)
                title = 'All Reports'
            
            # Generate file based on format
            try:
                if export_format == 'pdf':
                    file_content = self._generate_pdf(report_data, headers, title)
                    file_extension = 'pdf'
                    content_type = 'application/pdf'
                elif export_format == 'excel':
                    file_content = self._generate_excel(report_data, headers)
                    file_extension = 'xlsx'
                    content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                else:  # csv
                    file_content = self._generate_csv(report_data, headers)
                    file_extension = 'csv'
                    content_type = 'text/csv'
            except Exception as gen_error:
                print(f"❌ File generation error: {str(gen_error)}")
                raise
            
            # Create response
            response = HttpResponse(file_content, content_type=content_type)
            filename = f"report_{report_type}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{file_extension}"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response['Pragma'] = 'no-cache'
            response['Expires'] = '0'
            
            print(f"✅ Export success: {export_format} | size: {len(file_content)} bytes | file: {filename}")
            
            return response
            
        except Exception as e:
            print(f"❌ Export error: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)


class ApplyFiltersView(LoginRequiredMixin, View):
    """Apply filters and return filtered report data"""
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            category = data.get('category', 'all')
            date_from = data.get('date_from')
            date_to = data.get('date_to')
            export_format = data.get('export_format', 'pdf')
            
            # Build filter response
            filter_response = {
                'status': 'success',
                'message': 'Filters applied successfully',
                'filters_applied': {
                    'category': category,
                    'date_range': {
                        'from': date_from,
                        'to': date_to
                    },
                    'export_format': export_format
                },
                'records_found': 0
            }
            
            # Apply category filters
            if category == 'Sales Reports':
                filters = {}
                if date_from:
                    filters['order_date__gte'] = date_from
                if date_to:
                    filters['order_date__lte'] = date_to
                filter_response['records_found'] = Order.objects.filter(**filters).count()
            elif category == 'Inventory Reports':
                filter_response['records_found'] = InventoryLevel.objects.count()
            elif category == 'Customer Reports':
                filter_response['records_found'] = Customer.objects.count()
            elif category == 'Performance Reports':
                filter_response['records_found'] = Order.objects.count()
            else:
                filter_response['records_found'] = Order.objects.count()
            
            return JsonResponse(filter_response)
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=400)


# ============ Simple Views for URL placeholders ============

class StubView(TemplateView):
    template_name = 'stub.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['message'] = 'This feature is under development'
        return context


class SettingsView(LoginRequiredMixin, TemplateView):
    """System Settings Page - Configure SMTP, Security, etc."""
    template_name = 'settings.html'
    login_url = 'user:login'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from django.conf import settings
        
        context['page_title'] = 'System Settings'
        context['email_from'] = settings.DEFAULT_FROM_EMAIL
        context['email_host'] = settings.EMAIL_HOST
        context['email_port'] = settings.EMAIL_PORT
        context['user'] = self.request.user
        
        return context
    
    def dispatch(self, request, *args, **kwargs):
        # Only allow staff/admin users to access settings
        if not request.user.is_staff:
            messages.error(request, 'You do not have permission to access system settings.')
            return redirect('user:dashboard')
        return super().dispatch(request, *args, **kwargs)


class DashboardTestView(LoginRequiredMixin, View):
    """Dashboard Connection Test - Shows all backend data"""
    login_url = 'user:login'
    
    def get(self, request):
        # Calculate 30-day range
        today = timezone.now().date()
        thirty_days_ago = today - timedelta(days=30)
        
        # Get all orders from last 30 days
        orders_30days = Order.objects.filter(
            order_date__gte=timezone.make_aware(datetime.combine(thirty_days_ago, datetime.min.time()))
        )
        
        # 1. Revenue (30 Days)
        total_revenue = orders_30days.aggregate(Sum('total'))['total__sum'] or 0
        
        # 2. Total Orders
        total_orders = orders_30days.count()
        
        # 3. Average Order Value
        avg_order_value = orders_30days.aggregate(Avg('total'))['total__avg'] or 0
        
        # 4. Low Stock Items
        low_stock_items = InventoryLevel.objects.filter(
            quantity_on_hand__lt=10
        ).select_related('product', 'store').count()
        
        # 5. Top 5 Products
        top_products = OrderLine.objects.filter(
            order__order_date__gte=timezone.make_aware(datetime.combine(thirty_days_ago, datetime.min.time()))
        ).values('product__name', 'product__sku').annotate(
            total_sold=Sum('quantity'),
            total_revenue=Sum('line_total')
        ).order_by('-total_sold')[:5]
        
        # 6. Sales Trend (Daily sales for last 30 days)
        sales_trend = []
        for i in range(30):
            date = today - timedelta(days=i)
            daily_orders = Order.objects.filter(
                order_date__date=date
            )
            daily_revenue = daily_orders.aggregate(Sum('total'))['total__sum'] or 0
            sales_trend.append({
                'date': date.strftime('%Y-%m-%d'),
                'revenue': float(daily_revenue),
                'orders': daily_orders.count()
            })
        sales_trend.reverse()
        
        # 7. Inventory Status by Store
        inventory_by_store = InventoryLevel.objects.values('store__name').annotate(
            total_quantity=Sum('quantity_on_hand'),
            store_id=Sum('store__id')
        )
        
        # 8. Forecast data (next 7 days)
        forecast_data = []
        for i in range(7):
            date = today + timedelta(days=i)
            forecast_data.append({
                'date': date.strftime('%Y-%m-%d'),
                'forecasted_demand': 100 + (i * 15)
            })
        
        context = {
            'user': request.user,
            'total_revenue': f"${total_revenue:,.2f}",
            'total_orders': total_orders,
            'low_stock_items': low_stock_items,
            'avg_order_value': f"${avg_order_value:,.2f}",
            'top_products': list(top_products),
            'sales_trend': sales_trend,
            'inventory_by_store': list(inventory_by_store),
            'forecast_data': forecast_data,
        }
        
        return render(request, 'dashboard_test.html', context)


